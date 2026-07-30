"""
인플루언서 서칭 대시보드 (Streamlit 버전) - 전체 기능 반영판
"""

import io
import json
import re
from datetime import datetime, timedelta, timezone

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="인플루언서 서칭 대시보드", layout="wide")

st.markdown("""
<style>
    section[data-testid="stSidebar"] {
        background-color: #FFFFFF;
        border-right: 1px solid #E6E6EA;
    }
    section[data-testid="stSidebar"] label {
        font-size: 14px;
    }
    div[data-testid="stMetric"] {
        background-color: #FFFFFF;
        border: 1px solid #E6E6EA;
        border-radius: 12px;
        padding: 14px 16px;
    }
    div[data-testid="stMetricValue"] {
        font-size: 24px;
        font-weight: 700;
    }
    [data-testid="stVerticalBlockBorderWrapper"] {
        border-radius: 12px !important;
    }
    .stButton > button[kind="primary"] {
        background-color: #5B4FE0;
        border-color: #5B4FE0;
    }
    div[data-testid="stAlert"] {
        border-radius: 10px;
    }
    h1, h2, h3 {
        color: #18181B;
    }
</style>
""", unsafe_allow_html=True)

YOUTUBE_API_KEY = st.secrets.get("YOUTUBE_API_KEY", "")
YOUTUBE_BASE = "https://www.googleapis.com/youtube/v3"

CATEGORY_LIST = [
    "뷰티", "메이크업", "스킨케어", "패션", "데일리룩", "게임", "e스포츠",
    "요리", "베이킹", "여행", "캠핑", "육아", "반려동물", "운동/헬스",
]

CONTENT_TYPE_KEYWORDS = {
    "뷰티": ["뷰티", "화장품", "코스메틱", "뷰티템", "beauty"],
    "메이크업": ["메이크업", "화장법", "파운데이션", "아이섀도", "makeup"],
    "스킨케어": ["스킨케어", "피부관리", "클렌징", "세럼", "스킨"],
    "패션": ["패션", "스타일링", "코디", "ootd", "옷"],
    "데일리룩": ["데일리룩", "룩북", "데일리"],
    "브이로그": ["브이로그", "vlog", "일상"],
    "건강": ["건강", "웰빙", "웰에이징", "다이어트"],
    "리빙": ["리빙", "살림", "인테리어", "정리"],
    "푸드": ["푸드", "먹방", "맛집"],
    "게임": ["게임", "겜", "플레이", "gaming"],
    "e스포츠": ["e스포츠", "esports", "롤", "발로란트"],
    "요리": ["요리", "레시피", "쿠킹", "cooking"],
    "베이킹": ["베이킹", "제빵", "빵"],
    "여행": ["여행", "트래블", "travel"],
    "캠핑": ["캠핑", "차박", "camping"],
    "육아": ["육아", "아기", "육아템", "맘"],
    "반려동물": ["반려동물", "강아지", "고양이", "펫"],
    "운동/헬스": ["운동", "헬스", "홈트", "필라테스", "요가"],
}

AUDIENCE_KEYWORDS = [
    ("20대", ["20대", "이십대"]),
    ("30대", ["30대", "삼십대"]),
    ("40대 이상", ["40대", "50대", "중년", "웰에이징", "안티에이징"]),
    ("직장인", ["직장인", "회사원", "오피스"]),
    ("대학생", ["대학생", "캠퍼스"]),
    ("주부/엄마", ["주부", "엄마", "육아맘", "워킹맘"]),
    ("남성", ["남자", "남성", "그루밍"]),
    ("여성", ["여자", "여성", "언니", "누나"]),
    ("커플", ["커플", "연인"]),
]

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
INSTAGRAM_RE = re.compile(r"instagram\.com/([a-zA-Z0-9._]+)")
HANGUL_RE = re.compile(r"[\uac00-\ud7a3]")

AVATAR_PALETTE = [
    ("#EEECFC", "#3C3489"), ("#E1F5EE", "#085041"), ("#FBEAF0", "#72243E"),
    ("#FAEEDA", "#633806"), ("#E6F1FB", "#0C447C"), ("#EAF3DE", "#27500A"),
]

MIXSOON_CHANNELS = json.loads('''[{"no": 1, "name": "최종시안", "link": "https://www.youtube.com/channel/UC_1ETgrcrLjHOP0YJmkWLVQ", "subscribers": 362000, "content_type": "뷰티+브이로그", "features": "뷰티, 건강, 라이프스타일 팁을 공유하는 채널"}, {"no": 2, "name": "관리는 하고 살자", "link": "https://www.youtube.com/channel/UCVbW7vOtOnmjyDORZB4Sz0w", "subscribers": 358000, "content_type": "뷰티+건강", "features": "남성 뷰티, 건강, 자기관리 꿀팁 공유"}, {"no": 3, "name": "옌마드Yenmad", "link": "https://www.youtube.com/channel/UCAVIfs9U_yhPL234y2V2e4g", "subscribers": 355000, "content_type": "뷰티+브이로그", "features": "20대 자기관리, 뷰티, 라이프 브이로그"}, {"no": 4, "name": "뷰티숨BEAUTYSOOM", "link": "https://www.youtube.com/channel/UCLyq1gz11P-IYFmP093JaAw", "subscribers": 348000, "content_type": "뷰티+브이로그", "features": "전문가에게 배우는 ASMR 뷰티 팁 및 경험"}, {"no": 5, "name": "소의튜브soytube", "link": "https://www.youtube.com/channel/UC8qXNVuyfJLJtDf8vCr_xNw", "subscribers": 310000, "content_type": "뷰티", "features": "피부과 10년 경력의 미용 꿀팁과 제품 리뷰"}, {"no": 6, "name": "제이나 [Jaina]", "link": "https://www.youtube.com/channel/UCYsv-IHC-B-DiVMmJuqibcg", "subscribers": 281000, "content_type": "뷰티", "features": "화장품 마케터의 성분 분석, 비교 및 자사 제품 홍보"}, {"no": 7, "name": "쏭냥", "link": "https://www.youtube.com/channel/UC9P1tkf6e0hS7vzuDEuuAeA", "subscribers": 259000, "content_type": "뷰티+건강", "features": "맑은 피부를 위한 메이크업, 이너뷰티, 시술, 셀프케어 노하우를 공유하는 채널"}, {"no": 8, "name": "Coco Riley (코코 라일리)", "link": "https://www.youtube.com/channel/UCWQ0Xc5W-f2oSgx_Eohi9bQ", "subscribers": 258000, "content_type": "뷰티+건강", "features": "메이크업, 스킨케어 제품 리뷰와 피부 고민 해결 팁을 제공."}, {"no": 9, "name": "심톨 𝐒𝐈𝐌𝐓𝐎𝐇𝐋", "link": "https://www.youtube.com/channel/UCgRElRhf8PWv95nqExiQXHA", "subscribers": 249000, "content_type": "뷰티+패션", "features": "피부 시술, 패션 쇼핑 꿀팁 공유 채널"}, {"no": 10, "name": "단이 DANI", "link": "https://www.youtube.com/channel/UCxAdlKnB5u0a7A-iqRa5jYw", "subscribers": 235000, "content_type": "뷰티+패션", "features": "가성비와 꿀팁 위주 뷰티&패션 쇼핑 가이드"}, {"no": 11, "name": "임보라", "link": "https://www.youtube.com/channel/UCA1GLnIr2FL43PtWLmwhDvQ", "subscribers": 234000, "content_type": "뷰티+브이로그", "features": "임보라의 일상 브이로그 및 뷰티 콘텐츠"}, {"no": 12, "name": "bbomni 뽐니", "link": "https://www.youtube.com/channel/UCmPkZXhVfyLLWyHQM-l_KqA", "subscribers": 231000, "content_type": "뷰티", "features": "뷰티 신상 발굴과 전문가 꿀템을 공유하는 채널"}, {"no": 13, "name": "You need 윤이든", "link": "https://www.youtube.com/channel/UC57_jVZTM427hTqYMXf7utA", "subscribers": 216000, "content_type": "뷰티+패션", "features": "뷰티/패션/여행/다이어트 등 다양한 라이프스타일 공유"}, {"no": 14, "name": "톡신TOXIN", "link": "https://www.youtube.com/channel/UCHa6h8DGLYkAljbskQbIXHw", "subscribers": 206000, "content_type": "뷰티", "features": "무쌍 메이크업 & 스킨케어 꿀팁, 제품 리뷰 전문"}, {"no": 15, "name": "Kook연주", "link": "https://www.youtube.com/channel/UCSaEL-LEI41oepK5V5EXb2Q", "subscribers": 201000, "content_type": "뷰티+푸드", "features": "전직 화장품 연구원의 피부/음식 웰니스 채널"}, {"no": 16, "name": "오영주 OH!YOUNGJOO", "link": "https://www.youtube.com/channel/UCFsSyiKOtZD7YNt8uXHRlzg", "subscribers": 194000, "content_type": "브이로그+패션", "features": "오영주의 일상, 패션, 뷰티, 여행 정보를 담은 브이로그"}, {"no": 17, "name": "뭉컁 MungKyang", "link": "https://www.youtube.com/channel/UCDCvlVx1noYrb0rNN9N6Yqg", "subscribers": 188000, "content_type": "뷰티+브이로그", "features": "뷰티/생활 꿀템과 일상을 공유하는 채널"}, {"no": 18, "name": "유리하다", "link": "https://www.youtube.com/channel/UCgDgtCXMcNZ51QZ16K3M5cQ", "subscribers": 187000, "content_type": "뷰티+패션", "features": "뷰티 패션 쇼핑 리뷰를 유쾌하게 다루는 개그 채널"}, {"no": 19, "name": "화니 HWAN'E", "link": "https://www.youtube.com/channel/UCff7sQ_kjCEPZvr8h8US8ww", "subscribers": 172000, "content_type": "뷰티+브이로그", "features": "경험 기반 뷰티 꿀템 및 스킨케어 노하우 추천 채널"}, {"no": 20, "name": "이손 eson", "link": "https://www.youtube.com/channel/UCE5Q4Oy9GqzERUmI3E18s-A", "subscribers": 172000, "content_type": "뷰티+브이로그", "features": "뷰티 모델의 자기관리, 패션, 일상 브이로그"}, {"no": 21, "name": "림온 RIMON", "link": "https://www.youtube.com/channel/UCQG5MOnlAl5nUDR_JFOrrxA", "subscribers": 159000, "content_type": "뷰티+리빙", "features": "전직 뷰티BM의 관리 집착 피부/파데프리 꿀팁"}, {"no": 22, "name": "헤블 HEBLE", "link": "https://www.youtube.com/channel/UCb8_UiUdnM9S7r1aYPAsSMw", "subscribers": 159000, "content_type": "뷰티+리빙", "features": "피부 관리 팁과 실용적인 뷰티/리빙 꿀템 공유"}, {"no": 23, "name": "MysterLee터리", "link": "https://www.youtube.com/channel/UCm3aFTKTRsClNaN0rsYmQYg", "subscribers": 158000, "content_type": "뷰티+브이로그", "features": "센스있는 뷰티템 추천과 쇼핑 브이로그"}, {"no": 24, "name": "박비비 VIVI", "link": "https://www.youtube.com/channel/UCpSa5CzQedAxXFGfmeeFdIw", "subscribers": 138000, "content_type": "뷰티+브이로그", "features": "솔직한 팩트 기반 뷰티 리뷰 및 일상 브이로그"}, {"no": 25, "name": "래띠 LAETI", "link": "https://www.youtube.com/channel/UC4PAS3ck6trKI0RsM2Nskew", "subscribers": 134000, "content_type": "뷰티+브이로그", "features": "건강한 피부 관리와 일상을 담는 뷰티 유튜버"}, {"no": 26, "name": "한별두별[Hanna]", "link": "https://www.youtube.com/channel/UCYfAankzCbhjIV4EsU4Vzdw", "subscribers": 131000, "content_type": "뷰티", "features": "지성/트러블 피부를 위한 현실적 뷰티 꿀팁과 찐템 리뷰"}, {"no": 27, "name": "잠티", "link": "https://www.youtube.com/channel/UCxR4o2eiQz_nc2dUXGyysCw", "subscribers": 128000, "content_type": "뷰티", "features": "성분 기반 뷰티 제품 비판 및 분석 채널"}, {"no": 28, "name": "오드라이프oddlife", "link": "https://www.youtube.com/channel/UC9JU9mfYpg4KqxNZtD91GQg", "subscribers": 127000, "content_type": "뷰티+리빙", "features": "내돈내산 뷰티 및 리빙템 리뷰"}, {"no": 29, "name": "효블리 Hyovely", "link": "https://www.youtube.com/channel/UCB44fSGXvnBP6CVp10CenPg", "subscribers": 126000, "content_type": "뷰티+패션", "features": "봄웜 라이트의 신상 뷰티·패션 리뷰"}, {"no": 30, "name": "아우라M", "link": "https://www.youtube.com/channel/UCat2CSzaple02nnhbUSJ2zg", "subscribers": 125000, "content_type": "뷰티+패션", "features": "16년차 팩폭 에디터의 뷰티 패션 솔직 리뷰"}, {"no": 31, "name": "스칼렛 언니", "link": "https://www.youtube.com/channel/UCR8th9vv4cSnfoY-Gx5hnbA", "subscribers": 118000, "content_type": "뷰티+브이로그", "features": "화장품 개발자가 전하는 솔직한 뷰티 꿀팁"}, {"no": 32, "name": "하나둘세민 SEMIN", "link": "https://www.youtube.com/channel/UCr3rS9bIcUVSwnoCjV7yLPQ", "subscribers": 116000, "content_type": "뷰티+브이로그", "features": "다양한 메이크업과 일상을 공유하는 채널"}, {"no": 33, "name": "비트 BTE", "link": "https://www.youtube.com/channel/UCvGvf3Lk0Z5reU544g_SzPg", "subscribers": 111000, "content_type": "뷰티+리빙", "features": "피부 및 라이프스타일 관리 꿀팁과 추천템 리뷰 채널"}, {"no": 34, "name": "데일리 미라클", "link": "https://www.youtube.com/channel/UCk1BE73NfSUGz19LpNg9E-w", "subscribers": 382000, "content_type": "뷰티+패션", "features": "304050대 웰에이징 뷰티·패션·라이프스타일 꿀팁 채널(연령대 높음)"}, {"no": 35, "name": "뷰티마우스", "link": "https://www.youtube.com/channel/UCE7q5DzYcRYWObUNIpdO70A", "subscribers": 381000, "content_type": "뷰티+리빙", "features": "올영/쿠팡 중심의 뷰티 및 생활 꿀템 추천(연령대 높음)"}, {"no": 36, "name": "지인언니New Yorker Jiin", "link": "https://www.youtube.com/channel/UCWltt-7JK8k6IZjtuKzEdVA", "subscribers": 213000, "content_type": "뷰티+브이로그", "features": "50대 뉴요커가 전하는 뷰티, 건강, 라이프스타일 꿀팁(연령대 높음)"}, {"no": 37, "name": "모델엄마TV", "link": "https://www.youtube.com/channel/UCE2I3Ku0QqKRHg7kggyILMA", "subscribers": 192000, "content_type": "뷰티", "features": "동안 피부 위한 안티에이징 홈케어 뷰티 채널(연령대 높음)"}, {"no": 38, "name": "소살 Sora Salon", "link": "https://www.youtube.com/channel/UCPkgjWwYOiMKqNJpwdA8lPA", "subscribers": 135000, "content_type": "뷰티+패션", "features": "14년차 쇼호스트의 뷰티, 패션, 쇼핑 고급 정보(연령대 높음)"}, {"no": null, "name": "총 38개 · 발송 33개(회신 대기중) · 미발송 5개(연락처 없음) · 기준 2026-07-27 16:55", "link": null, "subscribers": null, "content_type": null, "features": null}, {"no": null, "name": "[2026-07-29_0956 기준] 회신 6 · 자동응답 0 · 반송 0 · 대기 27 · 미발송 5", "link": null, "subscribers": null, "content_type": null, "features": null}, {"no": null, "name": "[가격 출처] 2026-07-27~28 회신 메일 원문(Gmail) 그대로 표기 · 단위=만원 · 별도 표기 없으면 VAT 여부 원문에 명시 없음", "link": null, "subscribers": null, "content_type": null, "features": null}]''')


def avatar_color(name):
    if not name:
        return AVATAR_PALETTE[0]
    idx = sum(ord(c) for c in name) % len(AVATAR_PALETTE)
    return AVATAR_PALETTE[idx]


if "data" not in st.session_state:
    st.session_state.data = []
if "saved_lists" not in st.session_state:
    st.session_state.saved_lists = [
        {"name": "믹순 3차 — 기초화장품 채널", "created_at": "2026.07.29",
         "campaign_members": MIXSOON_CHANNELS},
    ]
if "history" not in st.session_state:
    st.session_state.history = []
if "nav_page" not in st.session_state:
    st.session_state.nav_page = "인플루언서 검색"
if "category_percents" not in st.session_state:
    st.session_state.category_percents = {}


def is_korean_channel(title, description, country, threshold=0.15):
    if country == "KR":
        return True
    text = f"{title} {description}"
    if not text.strip():
        return False
    hangul_count = len(HANGUL_RE.findall(text))
    return (hangul_count / max(len(text), 1)) >= threshold


def infer_content_type(title, description):
    text = f"{title or ''} {description or ''}".lower()
    scores = []
    for tag, keywords in CONTENT_TYPE_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw.lower() in text)
        if score > 0:
            scores.append((tag, score))
    scores.sort(key=lambda x: -x[1])
    top = [t for t, _ in scores[:2]]
    return "+".join(top) if top else None


def infer_audience(title, description):
    text = f"{title or ''} {description or ''}".lower()
    return [label for label, kws in AUDIENCE_KEYWORDS if any(kw.lower() in text for kw in kws)]


def subscriber_tier(subs):
    if subs is None:
        return ""
    if subs < 10_000:
        return "마이크로 인플루언서(1만 미만)"
    if subs < 100_000:
        return "미드티어(1만~10만)"
    if subs < 500_000:
        return "매크로(10만~50만)"
    return "메가 인플루언서(50만 이상)"


def generate_features(row):
    parts = []
    mix = row.get("category_mix")
    if mix:
        top = sorted(mix.items(), key=lambda x: -x[1])[:3]
        parts.append("최근 업로드는 " + ", ".join(f"{t} {p}%" for t, p in top) + " 비중으로 구성돼 있어요.")
    elif row.get("content_type"):
        parts.append(f"주로 {row['content_type'].replace('+', ', ')} 콘텐츠를 업로드해요.")
    else:
        parts.append("채널명/소개글만으로는 주력 콘텐츠를 특정하기 어려워요.")

    audience = infer_audience(row.get("title"), row.get("description"))
    if audience:
        parts.append(f"설명에 언급된 내용으로 볼 때 {', '.join(audience)} 타겟으로 보여요.")

    tier = subscriber_tier(row.get("subscriber_count"))
    if tier:
        parts.append(f"구독자 규모는 {tier}예요.")

    return " ".join(parts)


def is_healthy(subs, comment_reactivity):
    threshold = 0.5 if subs <= 10_000 else 0.25 if subs <= 100_000 else 0.1
    return comment_reactivity >= threshold


def is_recently_active(published_at, days=90):
    if not published_at:
        return False
    latest = datetime.fromisoformat(published_at.replace("Z", "+00:00"))
    return (datetime.now(timezone.utc) - latest) <= timedelta(days=days)


def channel_link(row):
    if row.get("platform") == "youtube":
        return f"https://www.youtube.com/channel/{row['channel_id']}"
    handle = row.get("instagram_handle")
    return f"https://www.instagram.com/{handle}/" if handle else None


def compute_cross_platform(data):
    ig_handles_from_yt = {
        d["instagram_handle"] for d in data
        if d.get("platform") == "youtube" and d.get("instagram_handle")
    }
    for d in data:
        if d.get("platform") == "youtube":
            d["cross_platform"] = bool(d.get("instagram_handle"))
        else:
            d["cross_platform"] = d.get("title") in ig_handles_from_yt


def balanced_slice(rows, limit):
    if not limit or len(rows) <= limit:
        return rows
    groups, order = {}, []
    for r in rows:
        key = r.get("category") or "기타"
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(r)
    picked, idx = [], 0
    while len(picked) < limit:
        added = False
        for key in order:
            if len(picked) >= limit:
                break
            g = groups[key]
            if idx < len(g):
                picked.append(g[idx])
                added = True
        if not added:
            break
        idx += 1
    return picked


def yt_get(endpoint, params):
    params = {**params, "key": YOUTUBE_API_KEY}
    resp = requests.get(f"{YOUTUBE_BASE}/{endpoint}", params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


def search_youtube(keyword, min_subs, max_subs, min_engagement, target_count, status=None):
    results = []
    page_token = None
    max_pages = 15

    for page in range(max_pages):
        if len(results) >= target_count:
            break
        if status:
            status.update(label=f'"{keyword}" 검색 중... ({len(results)}/{target_count}건 확보, {page + 1}페이지째)')

        params = {
            "q": keyword, "type": "channel", "part": "snippet", "maxResults": 50,
            "relevanceLanguage": "ko", "regionCode": "KR",
        }
        if page_token:
            params["pageToken"] = page_token

        search_data = yt_get("search", params)
        channel_ids = list({it["snippet"]["channelId"] for it in search_data.get("items", [])})
        page_token = search_data.get("nextPageToken")
        if not channel_ids:
            break

        channels_data = yt_get("channels", {
            "id": ",".join(channel_ids), "part": "snippet,statistics,contentDetails",
        })

        for item in channels_data.get("items", []):
            if len(results) >= target_count:
                break
            snippet = item["snippet"]
            description = snippet.get("description", "")
            subs = int(item["statistics"].get("subscriberCount", 0))
            country = snippet.get("country")

            if not (min_subs <= subs <= max_subs):
                continue
            if not is_korean_channel(snippet["title"], description, country):
                continue

            uploads_playlist = item["contentDetails"]["relatedPlaylists"]["uploads"]
            try:
                pl_data = yt_get("playlistItems", {
                    "playlistId": uploads_playlist, "part": "contentDetails", "maxResults": 8,
                })
                videos = [
                    {"videoId": v["contentDetails"]["videoId"],
                     "publishedAt": v["contentDetails"].get("videoPublishedAt")}
                    for v in pl_data.get("items", [])
                ]
            except requests.HTTPError:
                videos = []

            if not videos or not is_recently_active(videos[0]["publishedAt"]):
                continue

            comment_reactivity = 0.0
            category_mix = None

            videos_data = yt_get("videos", {
                "id": ",".join(v["videoId"] for v in videos), "part": "statistics,snippet",
            })
            items = videos_data.get("items", [])
            rates, tag_counts = [], {}
            for v in items:
                comments = int(v["statistics"].get("commentCount", 0))
                rates.append((comments / subs) * 100 if subs > 0 else 0)
                v_title = v.get("snippet", {}).get("title", "").lower()
                for tag, kws in CONTENT_TYPE_KEYWORDS.items():
                    if any(kw.lower() in v_title for kw in kws):
                        tag_counts[tag] = tag_counts.get(tag, 0) + 1
            if rates:
                comment_reactivity = round(sum(rates) / len(rates), 3)
            total_tagged = sum(tag_counts.values())
            if total_tagged:
                category_mix = {t: round(c / total_tagged * 100) for t, c in tag_counts.items()}
            recent_videos = [
                {
                    "videoId": v["id"],
                    "title": v.get("snippet", {}).get("title", ""),
                    "thumbnail": (v.get("snippet", {}).get("thumbnails", {}).get("medium", {}) or {}).get("url", ""),
                }
                for v in items[:4]
            ]

            if comment_reactivity < min_engagement:
                continue

            email_match = EMAIL_RE.search(description)
            ig_match = INSTAGRAM_RE.search(description)

            row = {
                "channel_id": item["id"],
                "platform": "youtube",
                "title": snippet["title"],
                "description": description,
                "thumbnail_url": (snippet.get("thumbnails", {}).get("medium", {}) or {}).get("url", ""),
                "country": country,
                "subscriber_count": subs,
                "view_count": int(item["statistics"].get("viewCount", 0)),
                "video_count": int(item["statistics"].get("videoCount", 0)),
                "engagement_rate_pct": comment_reactivity,
                "email": email_match.group(0) if email_match else None,
                "instagram_handle": ig_match.group(1) if ig_match else None,
                "keyword": keyword,
                "category": keyword,
                "category_mix": category_mix or {keyword: 100},
                "recent_videos": recent_videos,
                "searched_at": datetime.now(timezone.utc).isoformat(),
            }
            row["content_type"] = infer_content_type(row["title"], row["description"])
            row["features"] = generate_features(row)
            row["healthy_account"] = is_healthy(subs, comment_reactivity)
            results.append(row)

        if not page_token:
            break

    return results


def _base_sheet(wb, title):
    ws = wb.active
    ws.title = "인플루언서 리스트"
    ws.sheet_view.showGridLines = False
    thin_white = Side(style="thin", color="FFFFFFFF")
    ws.row_dimensions[2].height = 27.75
    ws["B2"] = title or "인플루언서 서칭 결과"
    ws["B2"].font = Font(name="맑은 고딕", size=14, bold=True, color="FF000000")
    ws["B2"].alignment = Alignment(vertical="center")
    ws["B2"].border = Border(top=thin_white, bottom=thin_white, left=thin_white, right=thin_white)
    return ws, thin_white


def export_to_excel(rows, title):
    wb = Workbook()
    ws, thin_white = _base_sheet(wb, title)
    thin_gray = Side(style="thin", color="FFA6A6A6")

    widths = [2.625, 31.375, 12.75, 10.75, 14.75, 19.0, 55.875, 16.5, 15.125, 28.75, 20.75, 10.75, 18.5, 8.875]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws["M2"] = datetime.now()
    ws["M2"].number_format = "mm-dd-yy"
    ws["M2"].font = Font(name="맑은 고딕", size=10)
    ws["M2"].border = Border(top=thin_white, bottom=thin_white, left=thin_white, right=thin_white)

    ws.row_dimensions[3].height = 23.25
    headers = ["이름", "플랫폼", "카테고리", "채널링크", "주력 컨텐츠 유형", "특징",
               "팔로워/구독자", "댓글 반응도(%)", "이메일", "인스타그램 핸들", "계정 상태", "두 플랫폼 동시 운영"]
    header_fill = PatternFill("solid", fgColor="FF002060")
    for i, h in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin_white, left=thin_white, right=thin_white)

    for r, row in enumerate(rows, start=4):
        link = channel_link(row)
        values = [
            row.get("title"),
            "유튜브" if row.get("platform") == "youtube" else "인스타그램",
            row.get("category"),
            link or "",
            row.get("content_type") or "",
            row.get("features") or "",
            row.get("subscriber_count"),
            row.get("engagement_rate_pct"),
            row.get("email") or "별도 확인이 필요함",
            row.get("instagram_handle") or "",
            "건강함" if row.get("healthy_account") else "저조함",
            "예" if row.get("cross_platform") else "아니오",
        ]
        for i, v in enumerate(values, start=2):
            cell = ws.cell(row=r, column=i, value=v)
            if i == 5 and link:
                cell.hyperlink = link
                cell.font = Font(name="맑은 고딕", size=11, color="FF0563C1", underline="single")
            else:
                cell.font = Font(name="맑은 고딕", size=11)
            cell.alignment = Alignment(horizontal="center", wrap_text=(i == 7))
            cell.border = Border(top=thin_gray, bottom=thin_gray, left=thin_gray, right=thin_gray)

    last_data_row = 3 + len(rows)
    for r in range(4, last_data_row + 1):
        ws.cell(row=r, column=1).border = Border(left=thin_white, top=thin_white, bottom=thin_white)
        ws.cell(row=r, column=14).border = Border(right=thin_white, top=thin_white, bottom=thin_white)
    close_row = last_data_row + 1
    for c in range(2, 15):
        ws.cell(row=close_row, column=c).border = Border(left=thin_white, right=thin_white, bottom=thin_white)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_campaign_excel(members, title):
    wb = Workbook()
    ws, thin_white = _base_sheet(wb, title)
    thin_gray = Side(style="thin", color="FFA6A6A6")

    widths = [2.625, 6, 31.375, 45, 12.75, 20, 45]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws["F2"] = datetime.now()
    ws["F2"].number_format = "mm-dd-yy"
    ws["F2"].font = Font(name="맑은 고딕", size=10)
    ws["F2"].border = Border(top=thin_white, bottom=thin_white, left=thin_white, right=thin_white)

    ws.row_dimensions[3].height = 23.25
    headers = ["번호", "채널명", "채널 링크", "구독자수", "주력 콘텐츠 유형", "특징"]
    header_fill = PatternFill("solid", fgColor="FF002060")
    for i, h in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin_white, left=thin_white, right=thin_white)

    for r, m in enumerate(members, start=4):
        values = [m.get("no"), m.get("name"), m.get("link"), m.get("subscribers"),
                  m.get("content_type"), m.get("features")]
        for i, v in enumerate(values, start=2):
            cell = ws.cell(row=r, column=i, value=v)
            if i == 4 and m.get("link"):
                cell.hyperlink = m["link"]
                cell.font = Font(name="맑은 고딕", size=11, color="FF0563C1", underline="single")
            else:
                cell.font = Font(name="맑은 고딕", size=11)
            cell.alignment = Alignment(horizontal="center", wrap_text=(i == 7))
            cell.border = Border(top=thin_gray, bottom=thin_gray, left=thin_gray, right=thin_gray)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@st.dialog("인플루언서 상세", width="large")
def show_detail(row):
    bg, fg = avatar_color(row["title"])
    initial = row["title"].strip()[:1] if row["title"] else "?"

    top_l, top_r = st.columns([1, 5])
    with top_l:
        st.markdown(
            f"""<div style="width:56px;height:56px;border-radius:50%;background:{bg};
            color:{fg};display:flex;align-items:center;justify-content:center;
            font-size:22px;font-weight:700;">{initial}</div>""",
            unsafe_allow_html=True,
        )
    with top_r:
        st.markdown(f"### {row['title']}")
        badges = [row["platform"] == "youtube" and "유튜브" or "인스타그램"]
        if row.get("cross_platform"):
            badges.append("유튜브+인스타")
        if row.get("category"):
            badges.append(row["category"])
        st.caption(" · ".join(badges))

    link_cols = st.columns(2)
    yt_link = channel_link(row) if row["platform"] == "youtube" else None
    ig_handle = row["title"] if row["platform"] == "instagram" else row.get("instagram_handle")
    if yt_link:
        link_cols[0].link_button("▶ 유튜브에서 보기", yt_link, use_container_width=True)
    if ig_handle:
        link_cols[1].link_button("📷 인스타그램에서 보기", f"https://www.instagram.com/{ig_handle}/", use_container_width=True)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("팔로워/구독자", f"{row['subscriber_count']:,}")
    m2.metric("댓글 반응도", f"{row['engagement_rate_pct']}%", help="최근 업로드 영상 최대 8개 평균 (댓글 수 ÷ 구독자 수 × 100)")
    m3.metric("업로드/게시물 수", f"{row.get('video_count', 0):,}개")
    m4.metric("누적 조회수", f"{row.get('view_count', 0):,}")

    healthy = row.get("healthy_account")
    if healthy:
        st.success("🟢 건강함 — 댓글 반응도가 팔로워 규모 대비 기준치 이상입니다")
    else:
        st.error("🔴 저조함 — 댓글 반응도가 팔로워 규모 대비 기준치보다 낮습니다")

    st.markdown("**연락처**")
    c1, c2 = st.columns(2)
    c1.text_input("이메일", value=row.get("email") or "별도 확인이 필요함", disabled=True, key="detail_email")
    c2.text_input("인스타그램", value=(f"@{row['instagram_handle']}" if row.get("instagram_handle") else "-"), disabled=True, key="detail_ig")

    recent = row.get("recent_videos") or []
    if recent:
        st.markdown("**최근 콘텐츠**")
        cols = st.columns(len(recent))
        for col, v in zip(cols, recent):
            with col:
                if v.get("thumbnail"):
                    st.image(v["thumbnail"], use_container_width=True)
                st.caption(f"[{v['title'][:22]}...](https://www.youtube.com/watch?v={v['videoId']})")

    st.divider()
    b1, b2, b3 = st.columns([2, 2, 1])
    with b1:
        with st.popover("☆ 리스트에 저장", use_container_width=True):
            name = st.text_input("리스트 이름", key="detail_save_name")
            if st.button("저장", key="detail_save_confirm"):
                final_name = name.strip() or f"리스트 {len(st.session_state.saved_lists) + 1}"
                st.session_state.saved_lists.insert(0, {
                    "name": final_name, "created_at": datetime.now().strftime("%Y.%m.%d"), "rows": [row],
                })
                st.success(f'"{final_name}"에 저장했어요.')
    with b2:
        if row.get("email"):
            subject = f"[협업 제안] {row['title']}님께"
            body = f"안녕하세요, {row['title']}님.%0D%0A%0D%0A협업 제안을 드리고 싶어 연락드립니다.%0D%0A%0D%0A감사합니다."
            st.link_button("이메일로 연락하기", f"mailto:{row['email']}?subject={subject}&body={body}", use_container_width=True)
        else:
            st.button("이메일로 연락하기 (이메일 없음)", disabled=True, use_container_width=True)
    with b3:
        if st.button("후보에서 제외", type="secondary"):
            st.session_state.data = [d for d in st.session_state.data if d["channel_id"] != row["channel_id"]]
            st.rerun()


with st.sidebar:
    st.markdown("## Influencer\nTools")
    st.session_state.nav_page = st.radio(
        "메뉴", ["인플루언서 검색", "저장한 리스트", "검색 히스토리"],
        label_visibility="collapsed",
        index=["인플루언서 검색", "저장한 리스트", "검색 히스토리"].index(st.session_state.nav_page),
    )

page = st.session_state.nav_page

if not YOUTUBE_API_KEY:
    st.warning(
        "⚠️ YouTube API 키가 설정되지 않았어요. `.streamlit/secrets.toml`에 "
        "`YOUTUBE_API_KEY = \"발급받은 키\"` 형식으로 넣어주세요. "
        "(Streamlit Cloud에 배포한 경우 앱 설정 > Secrets에서 등록)"
    )

if page == "인플루언서 검색":
    st.title("크리에이터 검색")
    st.caption(f"유튜브 · 인스타그램 인플루언서를 조건별로 검색하고 확인하세요 (현재 {len(st.session_state.data)}건 보유)")

    with st.container(border=True):
        row1 = st.columns([1.2, 1.6, 1.3, 1.3, 1, 1])
        with row1[0]:
            platforms = st.pills("플랫폼", ["유튜브", "인스타그램"], selection_mode="multi", default=["유튜브", "인스타그램"], key="f_platforms")
            st.caption("둘 다 선택 시 두 채널 동시 운영 계정만 표시")
        with row1[1]:
            categories = st.multiselect("카테고리", CATEGORY_LIST, key="f_categories")
            if categories:
                pc_cols = st.columns(len(categories))
                for col, cat in zip(pc_cols, categories):
                    with col:
                        val = st.number_input(f"{cat} 최소%", 0, 100, st.session_state.category_percents.get(cat, 0), step=5, key=f"pct_{cat}")
                        if val > 0:
                            st.session_state.category_percents[cat] = val
                        else:
                            st.session_state.category_percents.pop(cat, None)
        with row1[2]:
            sub_range = st.slider("팔로워/구독자", 0, 1_000_000, (0, 500_000), step=1000, key="f_subs")
        with row1[3]:
            min_engagement = st.slider(
                "최소 댓글 반응도(%)", 0.0, 6.0, 0.0, step=0.1, key="f_engagement",
                help="최근 업로드 영상 최대 8개의 평균 (댓글 수 ÷ 구독자 수 × 100). 좋아요는 포함하지 않습니다.",
            )
        with row1[4]:
            target_count = st.number_input("표시할 계정 수", min_value=1, value=20, step=5, key="f_count")
        with row1[5]:
            healthy_only = st.toggle("검증완료 계정만", key="f_healthy", help="댓글 반응도가 구독자 규모별 기준치(1만 이하 0.5%·10만 이하 0.25%·초과 0.1%) 이상인 계정만 표시")
            email_only = st.toggle("이메일 확인만", key="f_email", help="채널 설명란에 이메일이 적혀있는 계정만 표시")

        bcol1, bcol2 = st.columns([5, 1])
        run = bcol1.button("🔍 조회하기", type="primary", use_container_width=True)
        if bcol2.button("↺ 필터 초기화", use_container_width=True):
            for k in ["f_platforms", "f_categories", "f_subs", "f_engagement", "f_count", "f_healthy", "f_email"]:
                st.session_state.pop(k, None)
            st.session_state.category_percents = {}
            st.rerun()

    if run:
        if not categories:
            st.info("카테고리를 1개 이상 선택하면 유튜브에서 실시간으로 검색합니다. (지금은 보유 데이터만 필터링)")
        elif "유튜브" not in platforms:
            st.info("유튜브를 플랫폼에 포함해야 실시간 검색이 가능해요. (인스타그램 실시간 연동은 아직 준비 중입니다)")
        elif not YOUTUBE_API_KEY:
            st.error("YouTube API 키가 없어서 검색할 수 없어요. secrets 설정을 먼저 확인해주세요.")
        else:
            with st.status("실시간 검색 중...", expanded=True) as status:
                all_new = []
                for cat in categories:
                    new_rows = search_youtube(cat, sub_range[0], sub_range[1], min_engagement, target_count, status=status)
                    all_new.extend(new_rows)
                existing_ids = {r["channel_id"] for r in all_new}
                st.session_state.data = [r for r in st.session_state.data if r["channel_id"] not in existing_ids] + all_new
                compute_cross_platform(st.session_state.data)
                status.update(label=f"검색 완료 — {len(all_new)}건 추가 (전체 {len(st.session_state.data)}건)", state="complete")

            st.session_state.history.insert(0, {
                "categories": categories, "min_subs": sub_range[0], "max_subs": sub_range[1],
                "min_engagement": min_engagement, "result_count": len(all_new),
                "when": datetime.now().strftime("%Y.%m.%d %H:%M"),
            })

    rows = st.session_state.data
    plat_map = {"유튜브": "youtube", "인스타그램": "instagram"}
    selected_plats = {plat_map[p] for p in platforms}
    both_selected = selected_plats == {"youtube", "instagram"}
    if both_selected:
        rows = [r for r in rows if r.get("cross_platform")]
    else:
        rows = [r for r in rows if r["platform"] in selected_plats]

    percent_fallback = False
    if categories:
        percent_entries = {c: p for c, p in st.session_state.category_percents.items() if c in categories and p > 0}
        if percent_entries:
            def _pct_ok(r):
                mix = r.get("category_mix") or {}
                return all(mix.get(c, 0) >= p for c, p in percent_entries.items())
            strict = [r for r in rows if r["category"] in categories and _pct_ok(r)]
            if strict:
                rows = strict
            else:
                rows = [r for r in rows if r["category"] in categories]
                percent_fallback = True
        else:
            rows = [r for r in rows if r["category"] in categories]

    rows = [r for r in rows if sub_range[0] <= r["subscriber_count"] <= sub_range[1]]
    rows = [r for r in rows if r["engagement_rate_pct"] >= min_engagement]
    if healthy_only:
        rows = [r for r in rows if r.get("healthy_account")]
    if email_only:
        rows = [r for r in rows if r.get("email")]

    total_matched = len(rows)
    rows = sorted(rows, key=lambda r: r["engagement_rate_pct"], reverse=True)
    rows = balanced_slice(rows, target_count)

    if percent_fallback:
        st.warning(
            f"⚠️ 설정하신 비중 조건을 동시에 만족하는 채널은 없었어요. "
            f"대신 선택한 카테고리({', '.join(categories)})에 해당하는 채널을 모두 보여드립니다."
        )

    c1, c2 = st.columns(2)
    c1.metric("표시 중", f"{len(rows)}개", help=f"조건 충족 {total_matched}개 중")
    c2.metric("두 플랫폼 동시 운영", f"{sum(1 for r in rows if r.get('cross_platform'))}개",
              help="유튜브 채널 설명란과 인스타그램 프로필에 서로의 링크(또는 동일 계정명)가 함께 적혀있는 경우로 판별합니다.")

    col_a, col_b, col_c = st.columns([2, 1.4, 1.4])
    with col_a:
        list_name = st.text_input("리스트 이름", placeholder="리스트명 입력 (엑셀 제목/저장 리스트명으로 사용)", label_visibility="collapsed")
    with col_b:
        if st.button("☆ 리스트에 저장", disabled=not rows, use_container_width=True):
            name = list_name.strip() or f"리스트 {len(st.session_state.saved_lists) + 1}"
            st.session_state.saved_lists.insert(0, {
                "name": name, "created_at": datetime.now().strftime("%Y.%m.%d"), "rows": rows,
            })
            st.success(f'"{name}" 저장 완료! "저장한 리스트" 탭에서 확인하세요.')
    with col_c:
        if rows:
            excel_buf = export_to_excel(rows, list_name.strip() or "인플루언서 서칭 결과")
            st.download_button(
                "⬇ 엑셀로 내보내기", data=excel_buf,
                file_name=f"{(list_name.strip() or '인플루언서_서칭결과')}_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    if rows:
        df = pd.DataFrame([{
            "썸네일": r.get("thumbnail_url") or "",
            "이름": r["title"],
            "플랫폼": "유튜브" if r["platform"] == "youtube" else "인스타그램",
            "카테고리": r["category"],
            "채널링크": channel_link(r),
            "주력 콘텐츠 유형": r.get("content_type") or "-",
            "특징": r.get("features") or "-",
            "팔로워/구독자": r["subscriber_count"],
            "댓글 반응도(%)": r["engagement_rate_pct"],
            "이메일": r.get("email") or "별도 확인이 필요함",
            "계정 상태": "🟢 건강함" if r["healthy_account"] else "🔴 저조함",
        } for r in rows])

        event = st.dataframe(
            df, use_container_width=True, hide_index=True,
            column_config={
                "썸네일": st.column_config.ImageColumn("썸네일", width="small"),
                "채널링크": st.column_config.LinkColumn("채널링크", display_text="바로가기 ↗"),
                "댓글 반응도(%)": st.column_config.NumberColumn(
                    "댓글 반응도(%)", help="최근 업로드 영상 최대 8개 평균 (댓글 수 ÷ 구독자 수 × 100)"),
                "계정 상태": st.column_config.TextColumn(
                    "계정 상태", help="구독자 규모별 댓글 반응도 기준치 이상이면 건강함"),
            },
            on_select="rerun", selection_mode="single-row",
        )
        st.caption('이메일은 채널/프로필 설명에 공개된 경우만 표시됩니다. 표의 행을 선택하면 상세 정보가 팝업으로 뜹니다.')

        sel = event.selection.rows if event and event.selection else []
        if sel:
            show_detail(rows[sel[0]])
    else:
        st.info("조건에 맞는 인플루언서가 없습니다. 필터를 조정해보세요.")

elif page == "저장한 리스트":
    st.title("저장한 리스트")
    st.caption("캠페인별로 저장해둔 인플루언서 후보 리스트를 관리하세요")

    if not st.session_state.saved_lists:
        st.info("아직 저장한 리스트가 없어요. '인플루언서 검색' 탭에서 결과를 저장해보세요.")
    else:
        for i, lst in enumerate(st.session_state.saved_lists):
            is_campaign = "campaign_members" in lst
            count = len(lst.get("campaign_members") or lst.get("rows") or [])
            with st.container(border=True):
                col1, col2, col3 = st.columns([4, 1.2, 1])
                col1.markdown(f"**{lst['name']}**  \n채널 {count}개 · {lst['created_at']}" + ("  \n*(가격·연락처 등은 추가 예정)*" if is_campaign else ""))

                if is_campaign:
                    excel_buf = export_campaign_excel(lst["campaign_members"], lst["name"])
                else:
                    excel_buf = export_to_excel(lst["rows"], lst["name"])
                col2.download_button("⬇ 내보내기", data=excel_buf, file_name=f"{lst['name']}.xlsx",
                                      key=f"export_{i}", use_container_width=True)

                if col3.button("삭제", key=f"delete_{i}", use_container_width=True):
                    st.session_state.saved_lists.pop(i)
                    st.rerun()

                with st.expander("펼쳐보기"):
                    if is_campaign:
                        cdf = pd.DataFrame([{
                            "번호": m.get("no"), "채널명": m.get("name"), "채널링크": m.get("link"),
                            "구독자수": m.get("subscribers"), "주력 콘텐츠 유형": m.get("content_type"),
                            "특징": m.get("features"),
                        } for m in lst["campaign_members"]])
                        st.dataframe(cdf, use_container_width=True, hide_index=True,
                                     column_config={"채널링크": st.column_config.LinkColumn("채널링크", display_text="바로가기 ↗")})
                    else:
                        rdf = pd.DataFrame([{
                            "이름": r["title"], "플랫폼": "유튜브" if r["platform"] == "youtube" else "인스타그램",
                            "채널링크": channel_link(r), "팔로워/구독자": r["subscriber_count"],
                            "댓글 반응도(%)": r["engagement_rate_pct"],
                            "계정 상태": "🟢 건강함" if r["healthy_account"] else "🔴 저조함",
                        } for r in lst["rows"]])
                        st.dataframe(rdf, use_container_width=True, hide_index=True,
                                     column_config={"채널링크": st.column_config.LinkColumn("채널링크", display_text="바로가기 ↗")})

elif page == "검색 히스토리":
    st.title("검색 히스토리")
    st.caption("지금까지 실행한 검색 조건을 다시 불러올 수 있어요")

    if not st.session_state.history:
        st.info("아직 검색 기록이 없어요.")
    else:
        for i, h in enumerate(st.session_state.history):
            with st.container(border=True):
                c1, c2 = st.columns([5, 1])
                c1.markdown(
                    f"**{', '.join(h['categories'])}** · "
                    f"팔로워/구독자 {h['min_subs']:,}~{h['max_subs']:,} · "
                    f"최소 댓글 반응도 {h['min_engagement']}% · "
                    f"결과 {h['result_count']}건 · {h['when']}"
                )
                if c2.button("↺ 다시 검색", key=f"redo_{i}", use_container_width=True):
                    st.session_state.f_categories = h["categories"]
                    st.session_state.f_subs = (h["min_subs"], h["max_subs"])
                    st.session_state.f_engagement = h["min_engagement"]
                    st.session_state.nav_page = "인플루언서 검색"
                    st.rerun()
